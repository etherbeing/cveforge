import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def create_sales_report(filename="report.xlsx"):
    """
    Creates an Excel file with sales data and calculates total sales using a formula.

    Args:
        filename (str): The name of the Excel file to create.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sales Report"

    # Define headers
    headers = ["Product", "Quantity", "Unit Price", "Total Sale"]
    sheet.append(headers)

    # Sample data
    sales_data = [
        ("Laptop", 5, 1200),
        ("Mouse", 10, 25),
        ("Keyboard", 8, 75),
        ("Monitor", 3, 300),
    ]

    # Populate data and add formulas for "Total Sale"
    for row_idx, data in enumerate(sales_data, start=2): # Start from row 2 as row 1 is headers
        product, quantity, unit_price = data
        sheet.cell(row=row_idx, column=1, value=product)
        sheet.cell(row=row_idx, column=2, value=quantity)
        sheet.cell(row=row_idx, column=3, value=unit_price)
        
        # Add formula for Total Sale (D column): Quantity * Unit Price
        # Example: =B2*C2 for the first data row
        sheet.cell(row=row_idx, column=4, value=f"=B{row_idx}*C{row_idx}")

    # Add a row for total sales
    total_sales_row = len(sales_data) + 2 # +2 for header row and the current row
    sheet.cell(row=total_sales_row, column=1, value="Grand Total:")
    
    # Add formula for summing all "Total Sale" values
    # Example: =SUM(D2:D5)
    sheet.cell(row=total_sales_row, column=4, value=f"=SUM(D2:D{total_sales_row - 1})")

    # --- Styling ---
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15 # Set column width

    # Data row styling
    for row_idx in range(2, total_sales_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in [2, 3, 4]: # Quantity, Unit Price, Total Sale - center align
                cell.alignment = Alignment(horizontal="center")
            if col_idx == 4 and row_idx < total_sales_row: # Format "Total Sale" as currency
                cell.number_format = '"€"#,##0.00'
            elif col_idx == 4 and row_idx == total_sales_row: # Grand Total formatting
                cell.number_format = '"€"#,##0.00'
                cell.font = Font(bold=True)

    # Bold "Grand Total:" label
    sheet.cell(row=total_sales_row, column=1).font = Font(bold=True)
    sheet.cell(row=total_sales_row, column=1).alignment = Alignment(horizontal="right")
    
    # Save the workbook
    workbook.save(filename)
    print(f"Excel file '{filename}' created successfully with sales data and formulas.")

if __name__ == "__main__":
    create_sales_report()
